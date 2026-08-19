import time
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.webdriver.chrome.options import Options

options = Options()
options.binary_location = r"C:\Program Files\Google\Chrome\Application\chrome.exe"

try:
   driver = webdriver.Chrome(options=options)
   driver.get("https://botsdna.com/ServerAvailability/")
   driver.maximize_window()

except Exception as e:
   print(f"Error starting browser or loading page: {e}")
   exit()

time.sleep(2)

input_file = r"D:\Server Availability\input.xlsx"

try:
    workbook = load_workbook(filename=input_file)
except FileNotFoundError:
    print("Input File not found!")
    driver.quit()
    exit()

for sheet in workbook.worksheets:
    for index, row in enumerate(list(sheet.iter_rows(values_only=True))[1:],start=2):
    # for row in list(sheet.iter_rows(values_only=True))[1:]:
            if not any(row):
                break
    
            ROW = list(row)

            time.sleep(3)
            driver.get("https://botsdna.com/ServerAvailability/")
            time.sleep(2)

            username=ROW[0]
            pwd=ROW[1]
            server_name = str(ROW[2]) + "-" + str(ROW[3])
            SERVER_NAME=server_name.replace(".","-")

            print("Username:",username)
            print("Password:",pwd)
            print("Server name:",SERVER_NAME)

                            #Username
            try:
                driver.find_element(By.XPATH,"//input[@id='username']").send_keys(username)
                time.sleep(2)

            except Exception as e:
                print("Username field error:", e)
                print("Skipping this row...")
                continue

                            #Password
            try:
                driver.find_element(By.XPATH,"//input[@id='password']").send_keys(pwd)
                time.sleep(2)

            except Exception as e:
                print("Password field error:", e)
                print("Skipping this row...")
                continue

                           #Server name
            try:
                server= Select( driver.find_element( By.XPATH, "//select[@id='name']") )
                server.select_by_visible_text(SERVER_NAME)
                time.sleep(2)

            except Exception:
                print("Server name not found in dropdown list", SERVER_NAME)
                print("Skipping this row...")
                print()
                continue


            driver.find_element( By.XPATH, "//input[@value='Start Server']").click()
            time.sleep(2)

            status = driver.find_element(By.XPATH, "//div[@id='status']").text
            status = status.split("\n")[1]
            print("Status:", status)
            # print()

            sheet.cell(row=index,column=5).value=status

            try:
                workbook.save(input_file)
                # print()
                print("========================================")
                print("Excel updated successfully!")
                print("========================================")
                print()
            
            except PermissionError:
                print()
                print("Excel file is already open. Please close the Excel file and run the program again.")
                driver.quit()
                exit()

             
time.sleep(2)

try:
   driver.quit()
   print("Browser closed successfully.")

except Exception as e:
   print(f"Error closing browser: {e}")
