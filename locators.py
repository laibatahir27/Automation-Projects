from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
import time

options = Options()
options.binary_location = r"C:\Program Files\Google\Chrome\Application\chrome.exe"

try:
   driver = webdriver.Chrome(options=options)
   driver.get("https://botsdna.com/locator/")
   driver.maximize_window()

except Exception as e:
   print(f"Error starting browser or loading page: {e}")
   exit()


time.sleep(3)

input_file = r"C:\Users\Grace\Downloads\Locator-Output.xlsx"
output_file = r"C:\Users\Grace\Downloads\Locator-Output.xlsx"

try:
   workbook = load_workbook(filename=input_file)

except FileNotFoundError:
   print("Input file not found!")
   driver.quit()
   exit()

except Exception as e:
   print(f"Error opening Excel file: {e}")
   driver.quit()
   exit()


try:
   for sheet in workbook.worksheets:
      workbook.remove(sheet)

except Exception as e:
   print(f"Error removing old sheets: {e}")
   driver.quit()
   exit()


try:
   actual_table= driver.find_element(By.XPATH,"//table[not(contains(@style,'display: none;'))]")

except Exception as e:
   print(f"Table not found: {e}")
   driver.quit()
   exit()

try:
   headers = actual_table.find_elements(By.TAG_NAME, "th")
   rows = actual_table.find_elements(By.TAG_NAME, "tr")

except Exception as e:
   print(f"Error getting table data: {e}")
   driver.quit()
   exit()


if not headers:
   print("No headers found!")
   driver.quit()
   exit()

if not rows:
   print("No rows found!")
   driver.quit()
   exit()

print("Actual table's data:")
print(actual_table.text)

print("\nHeaders:")

for index, header in enumerate(headers[1:], start=1):
    try:
        country_name = header.text.strip()
        print(country_name)

        new_sheet = workbook.create_sheet(country_name)
        new_sheet.append(["CustomerName", "Number of Locations"])

    except Exception as e:
        print(f"Error creating sheet '{header.text}': {e}")
        continue

    for row in rows[1:]:
        try:
            print(row.text)

            datas = row.find_elements(By.TAG_NAME, "td")     
            customer_name = datas[0].text.strip()
            location = datas[index].text.strip()

            if location == "0":
                continue

            new_sheet.append([ customer_name,location])

        except IndexError as e:
            print(f"Index error while processing row: {e}")
            continue

        except Exception as e:
            print(f"Error processing row: {e}")
            continue

try:
   workbook.save(output_file)
   print("\nExcel file saved successfully!")

except PermissionError:
   print()
   print("Excel file is already open.")
   print("Please close the Excel file and run the program again.")
   driver.quit()
   exit()

except Exception as e:
   print(f"Error saving Excel file: {e}")
   driver.quit()
   exit()


time.sleep(2)

try:
   driver.quit()
   print("Browser closed successfully.")

except Exception as e:
   print(f"Error closing browser: {e}")