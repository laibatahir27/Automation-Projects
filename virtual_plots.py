from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import time

options = Options()
options.binary_location = r"C:\Program Files\Google\Chrome\Application\chrome.exe"

try:
   driver = webdriver.Chrome(options=options)
   driver.get("https://botsdna.com/vitrualplots/")
   driver.maximize_window()

except Exception as e:
   print(f"Error starting browser or loading page: {e}")
   exit()

time.sleep(2)

input_file = r"D:\Virtual Plots\input.xlsx"

try:
    workbook=load_workbook(filename=input_file)

except FileNotFoundError:
    print("Input File not found!")
    driver.quit()
    exit()

SMTP_SERVER = "smtp.gmail.com"
SMTP_PORT = 587
sender_email = "laibabaig1000@gmail.com"
receiver_email = "laibabaig1000@gmail.com"
app_password = "pgwl svmg lfyt jbcm"

for sheet in workbook.worksheets:
    for num, row in enumerate(list(sheet.iter_rows(values_only=True))[1:], start=2):
        ROW= list(row)
        col1= str(ROW[0]).replace("-", "")[-10:]
        col2= str(ROW[2]).replace("-", "")[-10:]
        plot_no= str(ROW[4])
        sqt= str(ROW[5])

        print(col1, col2)

        seller_found= False
        buyer_found= False
        seller_name= ""
        buyer_name= ""

        row_no= 2

        while True:
            try:
                table_value = driver.find_element(By.XPATH,f"html/body/center/table[1]/tbody/tr[{row_no}]/td[4]").text.strip()

                if table_value == col1:
                    driver.find_element(By.XPATH,f"html/body/center/table[1]/tbody/tr[{row_no}]/td[2]/input").click()
                    seller_name = driver.find_element(By.XPATH,f"html/body/center/table[1]/tbody/tr[{row_no}]/td[3]").text.strip()
                    seller_found = True
                    print("SELLER:", seller_name)
                    break

                row_no += 1

            except:
                print("SELLER NOT FOUND:", col1)
                break

        if not seller_found:
            continue

        row_no = 2

        while True:
            try:
                table_value = driver.find_element(By.XPATH,f"html/body/center/table[1]/tbody/tr[{row_no}]/td[4]").text.strip()

                if table_value == col2:
                    driver.find_element(By.XPATH,f"html/body/center/table[1]/tbody/tr[{row_no}]/td[1]/input").click()
                    buyer_name = driver.find_element(By.XPATH,f"html/body/center/table[1]/tbody/tr[{row_no}]/td[3]").text.strip()
                    buyer_found = True
                    print("BUYER:", buyer_name)
                    break

                row_no += 1

            except:
                print("BUYER NOT FOUND:", col2)
                break

        if not buyer_found:
            continue

        driver.find_element(By.XPATH,"html/body/center/table[2]/tbody/tr[1]/td/input").send_keys(plot_no)
        driver.find_element(By.XPATH,"html/body/center/table[2]/tbody/tr[2]/td/input").send_keys(sqt)
        driver.find_element(By.XPATH,"html/body/center/table[2]/tbody/tr[3]/td[2]/input").click()
        transac_no = driver.find_element(By.XPATH,"//p[@id='TransNo']").text
        print("Transaction Number:", transac_no)

        sheet.cell(row=num, column=7).value = transac_no
        workbook.save(input_file)

        subject= "Virtual Plot Booking Created Successfully"

        body= f"""Hello!

The virtual plot booking has been created successfully.

Seller Name: {seller_name}
Buyer Name: {buyer_name}
Plot Number: {plot_no}
Square Feet: {sqt}
Transaction Number: {transac_no}
"""

        msg = MIMEMultipart()
        msg["From"] = sender_email
        msg["To"] = receiver_email
        msg["Subject"] = subject
        msg.attach(MIMEText(body, "plain"))

        try:
            print("Sending email...")

            server = smtplib.SMTP(SMTP_SERVER, SMTP_PORT)
            server.starttls()
            server.login(sender_email, app_password)
            server.sendmail(sender_email, receiver_email, msg.as_string())
            server.quit()

            print("Email sent successfully!")

        except Exception as e:
            print("Error sending email:", e)

        time.sleep(2)

        driver.back()
        time.sleep(5)

        driver.find_element(By.XPATH,"html/body/center/table[2]/tbody/tr[1]/td/input").clear()
        driver.find_element(By.XPATH,"html/body/center/table[2]/tbody/tr[2]/td/input").clear()

        time.sleep(5)


time.sleep(2)

try:
   driver.quit()
   print("Browser closed successfully.")

except Exception as e:
   print(f"Error closing browser: {e}")

