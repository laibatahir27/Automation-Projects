from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
import time


workbook = load_workbook(filename=r'C:\Users\Grace\Downloads\output.xlsx')
sheet1 = workbook.active

options = Options()
options.binary_location = r"C:\Program Files\Google\Chrome\Application\chrome.exe"
driver=webdriver.Chrome(options=options)
driver.get("https://rpachallenge.com/")
driver.maximize_window()

time.sleep(2)

driver.find_element(By.XPATH, "//button[contains(text(),'Start')]").click()

time.sleep(2)

for row in list(sheet1.iter_rows(values_only=True))[1:]:
    
    driver.find_element(By.XPATH, "//input[@ng-reflect-name='labelFirstName']").send_keys(str(row[0]))
    driver.find_element(By.XPATH, "//input[@ng-reflect-name='labelLastName']").send_keys(str(row[1]))
    driver.find_element(By.XPATH, "//input[@ng-reflect-name='labelCompanyName']").send_keys(str(row[2]))
    driver.find_element(By.XPATH, "//input[@ng-reflect-name='labelRole']").send_keys(str(row[3]))
    driver.find_element(By.XPATH, "//input[@ng-reflect-name='labelAddress']").send_keys(str(row[4]))
    driver.find_element(By.XPATH, "//input[@ng-reflect-name='labelEmail']").send_keys(str(row[5]))
    driver.find_element(By.XPATH, "//input[@ng-reflect-name='labelPhone']").send_keys(str(row[6]))
    driver.find_element(By.XPATH, "//input[@type='submit']").click()
    
    time.sleep(1)

time.sleep(0.5)
driver.quit()