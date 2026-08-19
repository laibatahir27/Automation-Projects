from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.by import By
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import time

input_file = r"D:\server creation\input.xlsx"

options = Options()
options.binary_location = r"C:\Program Files\Google\Chrome\Application\chrome.exe"
try:
   driver = webdriver.Chrome(options=options)
   driver.get("https://botsdna.com/server/")
   driver.maximize_window()

except Exception as e:
   print(f"Error starting browser or loading page: {e}")
   exit()

time.sleep(2)

try:
    workbook=load_workbook(filename=input_file)

except FileNotFoundError:
    print("Input File not found!")
    driver.quit()
    exit()


hdd_index={"250 GB": 1,"500 GB": 2, "1 TB": 3,"2 TB": 4,"5 TB": 5}
app_index={"7-Zip": 1,"Adobe Acrobat Professional": 2,"Adobe Reader": 3,"Google Chrome": 4,"Microsoft Office": 5,"Microsoft SQL Server": 6,
             "MS Teams": 7,"Maxima": 8,"PuTTY": 9}


SMTP_SERVER = "smtp.gmail.com"
SMTP_PORT = 587
sender_email = "laibabaig1000@gmail.com"
receiver_email = "laibabaig1000@gmail.com"
app_password = "pgwl svmg lfyt jbcm"


for sheet in workbook.worksheets:

    for row in list(sheet.iter_rows(values_only=True))[1:]:
        if not any(row):
            break

        ROW = list(row)[1:5]
        print(ROW)

        a = str(ROW[0]).strip()
        b = str(ROW[1]).strip()
        c = str(ROW[2]).strip()
        d = str(ROW[3]).strip()

        apps = [app.strip() for app in d.split(",")]

        a_element = Select( driver.find_element( By.XPATH, "//select[@id='os']") )
        a_element.select_by_visible_text(a)
        time.sleep(1)


        b_element = Select( driver.find_element( By.XPATH,"//select[@id='Ram']" ) )
        b_element.select_by_visible_text(b)
        time.sleep(1)


        driver.find_element( By.XPATH,f"(//input[@id='hdd'])[{hdd_index[c]}]").click()
        time.sleep(1)


        for app in apps:
            driver.find_element( By.XPATH, f"(//input[@id='vehicle1'])[{app_index[app]}]").click()
            time.sleep(1)


        driver.find_element(By.XPATH,"//input[@id='CreateServer']").click()
        time.sleep(2)


        res1=driver.find_element(By.XPATH,"//div[@id='serverIP']/table/tbody/tr[1]/td[1]").text
        res2=driver.find_element(By.XPATH,"//div[@id='serverIP']/table/tbody/tr[1]/td[2]").text
        print(res1,":",res2)

        res3=driver.find_element(By.XPATH,"//div[@id='serverIP']/table/tbody/tr[2]/td[1]").text
        res4=driver.find_element(By.XPATH,"//div[@id='serverIP']/table/tbody/tr[2]/td[2]").text
        print(res3,":",res4)


        res5=driver.find_element(By.XPATH,"//div[@id='serverIP']/table/tbody/tr[3]/td[1]").text
        res6=driver.find_element(By.XPATH,"//div[@id='serverIP']/table/tbody/tr[3]/td[2]").text
        print(res5,":",res6)
        
        subject= "Server Created Successfully"

        body = f"""Hello!

The server has been created successfully.

{res1}: {res2}
{res3}: {res4}
{res5}: {res6}
"""

        msg= MIMEMultipart()

        msg["From"]= sender_email
        msg["To"]= receiver_email
        msg["Subject"]= subject
        msg.attach(MIMEText(body, "plain"))


        try:

            print("Sending email...")
            server = smtplib.SMTP(SMTP_SERVER,SMTP_PORT)
            server.starttls()
            server.login(sender_email,app_password )
            server.sendmail( sender_email,receiver_email,msg.as_string() )
            server.quit()
            print("Email sent successfully!")


        except Exception as e:
            print("Error sending email:", e)


        time.sleep(2)
        driver.back()
        time.sleep(2)


time.sleep(2)
try:
   driver.quit()
   print("Browser closed successfully.")

except Exception as e:
   print(f"Error closing browser: {e}")
