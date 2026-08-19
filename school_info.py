from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
import time

options = Options()
options.binary_location = r"C:\Program Files\Google\Chrome\Application\chrome.exe"

try:
   driver = webdriver.Chrome(options=options)
   driver.get("https://botsdna.com/school/")
   driver.maximize_window()

except Exception as e:
   print(f"Error starting browser or loading page: {e}")
   exit()

wait = WebDriverWait(driver, 20)

input_file = r"C:\Users\Grace\Downloads\Master Template.xlsx"
output_file = r"D:\School Info\Master Template.xlsx"

try:
    workbook=load_workbook(filename=input_file)

except FileNotFoundError:
    print("Input File not found!")
    driver.quit()
    exit()

for sheet in workbook.worksheets:
    for row_num, row in enumerate(list(sheet.iter_rows(values_only=True))[1:],start=2):
        ROW = list(row)
        print(ROW)

        time.sleep(3)
        driver.get("https://botsdna.com/school/")
        time.sleep(2)

        main_window = driver.current_window_handle

        driver.find_element(By.XPATH,"//input[@id='SchoolCode']").send_keys(ROW[0])

        time.sleep(2)

        driver.find_element(By.XPATH,"//input[@id='SearchSchool']").click()

        wait.until(lambda d: len(d.window_handles) > 1)

        for window in driver.window_handles:
            if window != main_window:
                driver.switch_to.window(window)
                break

        a = driver.find_element(By.XPATH,"/html/body/center/h1").text   # school name

        sheet.cell(row=row_num,column=2).value = a

        rows = driver.find_elements(By.XPATH,"//html/body/center/table/tbody/tr")

        for index, row in enumerate(rows, start=3): # start=3 bcoz first 2 columns already filled
            value = row.find_element(By.XPATH,"./td[2]").text
            sheet.cell(row=row_num,column=index).value = value

        time.sleep(5)

        print("Current window:", driver.current_window_handle)
        print("Current URL:", driver.current_url)

        driver.close()
        driver.switch_to.window(main_window)

        try:
            workbook.save(output_file)
            print()
            print("========================================")
            print("Excel updated successfully!")
            print("========================================")

        except PermissionError:
            print()
            print("Excel file is already open. Please close the Excel file and run the program again.")
            driver.quit()
            exit()

        

print("Workbook saved successfully.")

time.sleep(2)

try:
   driver.quit()
   print("Browser closed successfully.")

except Exception as e:
   print(f"Error closing browser: {e}")
