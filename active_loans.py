from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
import time
import zipfile
import os

input_file = r"C:\Users\Grace\Downloads\input (1).xlsx"

options = Options()
options.binary_location = r"C:\Program Files\Google\Chrome\Application\chrome.exe"
try:
   driver = webdriver.Chrome(options=options)
   driver.get("https://botsdna.com/ActiveLoans/")
   driver.maximize_window()

except Exception as e:
   print(f"Error starting browser or loading page: {e}")
   exit()

time.sleep(2)

try:
    table = driver.find_element(By.XPATH, "//table")
    print("Table's data:")
    print(table.text)

except Exception as e:
   print(f"Table not found: {e}")
   driver.quit()
   exit()

try:
    workbook = load_workbook(filename=input_file)

except FileNotFoundError:
    print("Input File not found!")
    driver.quit()
    exit()


for sheet in workbook.worksheets:
    for row_num, row in enumerate(list(sheet.iter_rows(values_only=True))[1:],start=2):
        try:
            account_number = str(row[0])

            print()
            print("========================================")
            print("Excel Row:", row_num)
            print("Account Number:", account_number)
            print("========================================")

            last_four = account_number[-4:]

            print("Last four digits of excel:", last_four)

            loan_link = driver.find_element(By.XPATH,f"//html/body/center/table/tbody/tr/td[2]/a[contains(text(),'-{last_four}')]" )
            data= loan_link.text

            print("Loan Code:", data)
            DATA = data.split("-")
            print("Extracted data:", DATA[1])


            website_row= loan_link.find_element(By.XPATH,"./ancestor::tr")

            status = website_row.find_element(By.XPATH,"./td[1]").text

            print("Status:", status)

            pan = website_row.find_element(By.XPATH,"./td[3]").text

            print("PAN Number:", pan)

            sheet.cell(row=row_num,column=7).value = pan

            sheet.cell(row=row_num,column=8).value = status

            loan_link.click()

            time.sleep(2)

            path = rf"C:\Users\Grace\Downloads\{account_number}.zip"

            print("Complete Path:", path)

            if not os.path.exists(path):
                print("ZIP file not found!")
                continue


            print("ZIP file found!")

            with zipfile.ZipFile(path, "r") as zip_file:

                files = zip_file.namelist()

                print("Files inside ZIP:", files)

                txt_file = None

                for file in files:

                    if file.lower().endswith(".txt"):

                        txt_file = file

                        break



                if txt_file is None:
                    print("TXT file not found!")
                    continue


                print("TXT file found:", txt_file)
                text_data = zip_file.read(txt_file).decode("utf-8")

                for line in text_data.splitlines():
                    line = line.strip()

                    if line.startswith("Bank:"):
                        bank= line.replace("Bank:", "").strip()
                        sheet.cell(row=row_num,column=2).value = bank
                        print("Bank:", bank)


                    elif line.startswith("Branch:"):
                        branch= line.replace("Branch:", "").strip()
                        sheet.cell(row=row_num,column=3).value=branch
                        print("Branch:",branch)


                    elif line.startswith("Loan Taken On:"):
                        loan_taken_on= line.replace("Loan Taken On:", "").strip()
                        sheet.cell(row=row_num,column=4).value=loan_taken_on
                        print("Loan Taken On:",loan_taken_on)


                    elif line.startswith("Amount:"):
                        amount= line.replace("Amount:", "").strip()
                        sheet.cell(row=row_num,column=5).value=amount
                        print("Amount:",amount)


                    elif line.startswith("EMI(month):"):
                        emi= line.replace("EMI(month):", "").strip()
                        sheet.cell(row=row_num,column=6).value=emi
                        print("EMI(month):",emi)
                        
            try:
                workbook.save(input_file)
                print()
                print("========================================")
                print("Excel updated successfully!")
                print("========================================")
            except PermissionError:
                print()
                print("Excel file is already open. Please close the Excel file and run the program again.")
                driver.quit()
                exit()
                


        except Exception as e:
            print("Error:",e)
            continue


time.sleep(2)
try:
   driver.quit()
   print("Browser closed successfully.")

except Exception as e:
   print(f"Error closing browser: {e}")