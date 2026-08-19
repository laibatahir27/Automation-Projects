import time
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options

options = Options()
options.binary_location = r"C:\Program Files\Google\Chrome\Application\chrome.exe"

driver = webdriver.Chrome(options=options)

driver.get("https://www.theautomationchallenge.com/")
driver.maximize_window()

time.sleep(1)

driver.find_element(By.XPATH,"//button[@tabindex='14']").click()
time.sleep(1)

driver.find_element(By.XPATH,"//button[starts-with(text(),'OR LOGIN')]").click()
time.sleep(1)

driver.find_element(By.XPATH,"(//input[@placeholder='Email'])[1]").send_keys("laibatahir1000@gmail.com")
time.sleep(1)

driver.find_element(By.XPATH,"//input[@placeholder='Password'][1]").send_keys("pfdgp108")
time.sleep(1)

driver.find_element(By.XPATH,"//button[text()='LOG IN']").click()
time.sleep(1)

driver.find_element(By.XPATH,"//*[text()='Start']").click()
time.sleep(5)


input_file = r"D:\challenge.xlsx"

try:
    workbook = load_workbook(filename=input_file)
except FileNotFoundError:
    print("Input File not found!")
    driver.quit()
    exit()

for sheet in workbook.worksheets:
    #for index, row in enumerate(list(sheet.iter_rows(values_only=True))[1:],start=2):
    for row in list(sheet.iter_rows(values_only=True))[1:]:
            if not any(row):
                break
    
            ROW = list(row)

            e_id=ROW[0]
            company_name=ROW[1]
            sector=ROW[2]
            address=ROW[3]
            tool=ROW[4]
            saving=ROW[5]
            date=ROW[6]

            print("Employee ID:",e_id)
            print("Company Name:",company_name)
            print("Sector:",sector)
            print("Address:",address)
            print("Tool:",tool)
            print("Saving:",saving)
            print("Date:",date)

            # driver.find_element(By.XPATH,"//*[text()='Start']").click()
            # time.sleep(5)

            driver.find_element(By.XPATH,"//*[contains(text(),'Company Name')][not(ancestor::div[contains(@style,'display: none')])]/ancestor::div[contains(@class,'bubble-element Group')][1]//input").send_keys(company_name)
            time.sleep(3)

            driver.find_element(By.XPATH,"//*[contains(text(),'Address')][not(ancestor::div[contains(@style,'display: none')])]/ancestor::div[contains(@class,'bubble-element Group')][1]//input").send_keys(address)
            time.sleep(3)

            driver.find_element(By.XPATH,"//*[contains(text(),'EIN')][not(ancestor::div[contains(@style,'display: none')])]/ancestor::div[contains(@class,'bubble-element Group')][1]//input").send_keys(e_id)
            time.sleep(3)

            driver.find_element(By.XPATH,"//*[contains(text(),'Sector')][not(ancestor::div[contains(@style,'display: none')])]/ancestor::div[contains(@class,'bubble-element Group')][1]//input").send_keys(sector)
            time.sleep(3)

            driver.find_element(By.XPATH,"//*[contains(text(),'Automation Tool')][not(ancestor::div[contains(@style,'display: none')])]/ancestor::div[contains(@class,'bubble-element Group')][1]//input").send_keys(tool)
            time.sleep(3)

            driver.find_element(By.XPATH,"//*[contains(text(),'Annual Saving')][not(ancestor::div[contains(@style,'display: none')])]/ancestor::div[contains(@class,'bubble-element Group')][1]//input").send_keys(saving)
            time.sleep(3)

            driver.find_element(By.XPATH,"//*[contains(text(),'Date')][not(ancestor::div[contains(@style,'display: none')])]/ancestor::div[contains(@class,'bubble-element Group')][1]//input").send_keys(date)
            time.sleep(3)

            driver.find_element(By.XPATH,"//*[contains(text(),'Submit')][not(ancestor::div[contains(@style,'display: none')])]").click()
            time.sleep(3)


            time.sleep(10)


driver.quit()

            