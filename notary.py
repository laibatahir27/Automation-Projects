from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.by import By
import time

options=Options()
options.binary_location=r"C:\Program Files\Google\Chrome\Application\chrome.exe"
try:
   driver = webdriver.Chrome(options=options)
   driver.get("https://botsdna.com/notaries/")
   driver.maximize_window()

except Exception as e:
   print(f"Error starting browser or loading page: {e}")
   exit()

time.sleep(2)

input_file= r"C:\Users\Grace\Downloads\AP-ADVOCATES.xlsx"
output_file = r"D:\Notary\AP-ADVOCATES.xlsx"


try:
    workbook=load_workbook(filename=input_file)

except FileNotFoundError:
    print("Input File not found!")
    driver.quit()
    exit()

for sheet in workbook.worksheets:    
    district = None

    for index,row in enumerate(list(sheet.iter_rows(values_only=True))[1:],start=2):
            
            rows=list(row)
            print(rows)
            if rows[1] is not None and rows[2] is not None:
                  
                driver.find_element(By.XPATH,"//input[@id='notary']").send_keys(str(rows[1]))
                driver.find_element(By.XPATH,"//input[@id='area']").send_keys(str(rows[2]))
                print(district)
                ddelement= Select(driver.find_element(By.ID,"DIST"))
                ddelement.select_by_visible_text(district)
                time.sleep(3)
                driver.find_element(By.XPATH,"//input[@value='Submit Notary']").click()
                time.sleep(2)
                
                transac=driver.find_element(By.XPATH,"//p[@id='TransNo']").text
                print(transac)
                sheet.cell(row=index,column=4).value=transac

                try:
                    workbook.save(output_file)
                    print()
                    print("========================================")
                    print("Excel updated successfully!")
                    print("========================================")
                except PermissionError:
                    print()
                    print("Excel file is already open. Please close the Excel file and run the program again.")
                    driver.quit()
                    exit()
                                
                

                time.sleep(2)
                driver.back()
            else:
                district= rows[0].replace(" DIST", "").strip()


                                 
time.sleep(2)

try:
   driver.quit()
   print("Browser closed successfully.")

except Exception as e:
   print(f"Error closing browser: {e}")
